"""
Define option and translation settings and prepare dataset for option calculation.
"""


import datetime
import pandas as pd
import re
from openpyxl import load_workbook
from pandas.api.types import is_string_dtype, is_object_dtype

from fermatrica_utils import StableClass
from fermatrica.basics.basics import fermatrica_error


def trans_dict_create(dt_trans: pd.DataFrame | None = None
                      , path: str | None = None
                      , sheet: str | None = 'translation') -> dict:
    """
    Create translation rules dictionary from XLSX path or y from DataFrame. Set either `path` or
    `dt_trans` respectively. It binds tool name with translation function, data variable and max
    budget if curves to be calculated.

    Translation dictionary is vital for option calculation and hence for the export of future
    periods.

    :param dt_trans: translation dataframe (as `translation` sheet in "options.xlsx" file)
    :param path: path to XLSX file like "options.xlsx" with `sheet` sheet
    :param sheet: name of the sheet with translation data (only if `path` is set), defaults to
        "translation"
    :return: translation dictionary
    """

    # load

    if not pd.isna(path) and path != "":

        wb = load_workbook(path)
        shts = wb.sheetnames
        wb.close()

        if sheet not in shts:
            fermatrica_error('Error reading XLSX file with translation rules: ' + path + '. Sheet ' + sheet + ' not found')

        try:
            dt_trans = pd.read_excel(path, sheet_name=sheet)
        except Exception as e:
            fermatrica_error('Error reading XLSX file with translation rules: ' + path + '. File not found')

    elif pd.isna(dt_trans):
        fermatrica_error('Error creating translation rules dictionary. Not dict nor path to XLSX file is provided')

    # convert to dictionary

    dt_trans = dt_trans.loc[(dt_trans['function'] != '') & (pd.notna(dt_trans['function'])), :]

    dt_trans.set_index('item', inplace=True)
    dct = dt_trans.to_dict(orient='dict')

    return dct


def budget_dict_create(dct: dict | pd.DataFrame | None = None
                       , path: str | None = None
                       , sheet: str | None = 'main'
                       , vat_rate: float = 0.) -> dict:
    """
    Create budget dictionary from raw dictionary or from XLSX path. Budget dictionary is dictionary
    containing named options prepared by user by hand.

    Use the function to load manually prepared options to the environment.
    Set either `dct` to create from existing dictionary or `dct` to load from disc.

    :param dct: raw dictionary (in a form of a dictionary or a dataframe)
    :param path: path to XLSX file like "options.xlsx" with `sheet` sheet
    :param sheet: name of the sheet with option data (only if `path` is set), defaults to
        "main"
    :param vat_rate: VAT rate, defaults to 0. Set it different if costs used in the model include VAT
    :return: dictionary with named options as separate items
    """

    # load

    if not pd.isna(path) and path != "":

        wb = load_workbook(path)
        shts = wb.sheetnames
        wb.close()

        if sheet not in shts:
            fermatrica_error('Error reading XLSX file with budget options: ' + path + '. Sheet ' + sheet + ' not found')

        try:
            tmp = pd.read_excel(path, sheet_name=sheet)
        except Exception as e:
            fermatrica_error('Error reading XLSX file with budget options: ' + path + '. File not found')

        tmp = tmp.convert_dtypes(infer_objects=True, convert_string=True, convert_integer=True
                                 , convert_boolean=False, convert_floating=True)
        tmp['option'] = tmp['option'].astype('string')

        tmp.set_index('option', inplace=True)
        dct = tmp.to_dict(orient='index')

    elif isinstance(dct, pd.DataFrame):

        dct = dct.convert_dtypes(infer_objects=True, convert_string=True, convert_integer=True
                                 , convert_boolean=False, convert_floating=True)
        dct['option'] = dct['option'].astype('string')

        dct.set_index('option', inplace=True)
        dct = dct.to_dict(orient='index')

    elif pd.isna(dct):
        fermatrica_error('Error creating budget options dictionary. Not dict nor path to XLSX file is provided')

    # apply VAT & calculate total budget

    for k, v in dct.items():
        bdg = 0
        for k1 in v:
            if re.match(r'bdg_', k1):
                if vat_rate > 0:
                    v[k1] *= (1 + vat_rate)
                bdg += v[k1]
        v['bdg'] = bdg

    return dct


def media_ssn_apply(ds: pd.DataFrame
                    , dt_ssn: pd.DataFrame | None = None
                    , path: str | None = None
                    , sheet: str | None = 'data') -> pd.DataFrame:
    """
    Apply media seasonality to main dataset. Media seasonality is here for distribution of
    yearly budget by specific periods (months, weeks) within year.

    Media seasonality data to be passed either via `path` and `sheet` args (to load from
    disc) or as already existing dataframe as `dt_ssn` argument.

    :param ds: main dataset
    :param dt_ssn: media seasonal distribution data (optional)
    :param path: path to XLSX file like "seasonality.xlsx" with `sheet` sheet
    :param sheet: name of the sheet with option data (only if `path` is set), defaults to
        "data"
    :return: `ds` main dataset
    """

    # load

    if not pd.isna(path) and path != "":

        wb = load_workbook(path)
        shts = wb.sheetnames
        wb.close()

        if sheet not in shts:
            fermatrica_error('Error reading XLSX file with media seasonality: ' + path + '. Sheet ' + sheet + ' not found')

        try:
            dt_ssn = pd.read_excel(path, sheet_name=sheet)
        except Exception as e:
            fermatrica_error('Error reading XLSX file with media seasonality: ' + path + '. File not found')

    elif dt_ssn is None:
        fermatrica_error('Error creating media seasonality frame. Not frame nor path to XLSX file is provided')

    # remove check row

    period_var = dt_ssn.columns.tolist()[0]

    if is_string_dtype(dt_ssn[period_var]) | is_object_dtype(dt_ssn[period_var]):
        dt_ssn = dt_ssn[dt_ssn[period_var].str.lower() != 'total']

    # rename seasonality columns and cleanse main dataset

    dt_ssn.columns = ['opt_media_ssn_' + x if x != period_var else x for x in dt_ssn.columns.tolist()]

    for col in dt_ssn.columns.tolist():
        if col in ds.columns:
            del ds[col]

    # cast types and prepare merge

    if period_var == 'month':
        dt_ssn[period_var] = pd.to_numeric(dt_ssn.loc[:, period_var])
        ds[period_var] = ds.loc[:, 'date'].dt.month
    elif period_var == 'week_number':
        dt_ssn[period_var] = pd.to_numeric(dt_ssn.loc[:, period_var])
        ds[period_var] = ds.loc[:, 'date'].dt.isocalendar().week
    elif period_var == 'day':
        dt_ssn[period_var] = pd.to_datetime(dt_ssn.loc[:, period_var], format="%Y-%m-%d")
        ds[period_var] = ds.loc[:, 'day']

    # merge with data

    ds = ds.merge(dt_ssn, how='left', on=period_var, sort=False, copy=False)

    return ds


class OptionSettings(StableClass):
    """
    Option settings / controls. Effectively structure rather than class.
    """

    target: list
    date_start: str | datetime.date
    date_end: str | datetime.date
    ref_date_start: str | datetime.date
    ref_date_end: str | datetime.date
    plan_period: str  # 'exact', 'year', 'hy', 'quarter'
    apply_vars: list

    def __init__(
            self
            , target: list | tuple | str
            , date_start: str | datetime.date
            , date_end: str | datetime.date
            , ref_date_start: str | datetime.date
            , ref_date_end: str | datetime.date
            , apply_vars: list | tuple | str = ('superbrand',)
            , plan_period: str = 'year'
    ):
        """
        Initialize class.

        :param target: values of variables listed in `apply_vars` to apply option. Could be string
            "my_brand" or tuple / list of strings or of tuples / lists [["my_brand"], ["market_1", "market_2"]]
        :param date_start: start of the period to apply option
        :param date_end: end of the period to apply option
        :param ref_date_start: start of the reference period (to calculate growth)
        :param ref_date_end: end of the reference period (to calculate growth)
        :param apply_vars: variable to apply option. Could be string
            "superbrand" or tuple / list of strings ["superbrand", "market"]
        :param plan_period: planning period: 'exact', 'year', 'hy', 'quarter'
        """

        if isinstance(target, str):
            target = [target]
        if isinstance(apply_vars, str):
            apply_vars = [apply_vars]

        self.target = target
        self.date_start = pd.to_datetime(date_start)
        self.date_end = pd.to_datetime(date_end)
        self.ref_date_start = pd.to_datetime(ref_date_start)
        self.ref_date_end = pd.to_datetime(ref_date_end)
        self.plan_period = plan_period
        self.apply_vars = apply_vars
        self.zip_mask = None
